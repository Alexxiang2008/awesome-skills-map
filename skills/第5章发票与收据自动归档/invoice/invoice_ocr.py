#!/usr/bin/env python3
"""
发票自动归档工具（简化版）
使用百度OCR API识别增值税发票，自动分类汇总生成Excel报表
"""

import os
import sys
import json
import base64
import argparse
from pathlib import Path
from datetime import datetime
from collections import defaultdict

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# PDF处理（可选）
try:
    from pdf2image import convert_from_path
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False

# 配置文件路径
CONFIG_PATH = Path.home() / ".invoice-ocr-config.json"

# 发票分类规则
CATEGORY_RULES = {
    "餐饮": ["餐", "饮", "食品", "饭", "酒", "茶", "咖啡", "外卖", "美团", "饿了么"],
    "交通": ["打车", "出租", "滴滴", "网约车", "加油", "停车", "高速", "过路", "汽油", "柴油", "车费", "交通", "运输", "地铁", "客运", "铁路"],
    "办公": ["办公", "文具", "打印", "复印", "纸", "笔", "耗材", "设备"],
    "住宿": ["住宿", "酒店", "宾馆", "旅馆", "客房"],
    "通讯": ["话费", "流量", "通讯", "电话", "宽带", "网络"],
    "服务": ["服务费", "咨询", "技术服务", "软件", "会员"],
}


def load_config():
    """加载配置文件"""
    if not CONFIG_PATH.exists():
        print(f"错误：配置文件不存在，请创建 {CONFIG_PATH}")
        print("配置文件格式：")
        print(json.dumps({"api_key": "你的API Key", "secret_key": "你的Secret Key"}, indent=2, ensure_ascii=False))
        sys.exit(1)

    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        config = json.load(f)

    if "api_key" not in config or "secret_key" not in config:
        print("错误：配置文件缺少 api_key 或 secret_key")
        sys.exit(1)

    return config


def get_access_token(api_key: str, secret_key: str) -> str:
    """获取百度API访问令牌"""
    url = "https://aip.baidubce.com/oauth/2.0/token"
    params = {
        "grant_type": "client_credentials",
        "client_id": api_key,
        "client_secret": secret_key
    }
    response = requests.post(url, params=params)
    result = response.json()

    if "access_token" not in result:
        print(f"获取access_token失败: {result}")
        sys.exit(1)

    return result["access_token"]


def image_to_base64(image_path: str) -> str:
    """将图片转换为base64编码"""
    with open(image_path, "rb") as f:
        return base64.b64encode(f.read()).decode("utf-8")


def pdf_to_images(pdf_path: str, output_dir: str) -> list:
    """将PDF转换为图片列表"""
    if not PDF_SUPPORT:
        print(f"警告：未安装pdf2image，跳过PDF文件: {pdf_path}")
        return []

    images = convert_from_path(pdf_path)
    image_paths = []

    for i, image in enumerate(images):
        image_path = os.path.join(output_dir, f"temp_pdf_page_{i}.png")
        image.save(image_path, "PNG")
        image_paths.append(image_path)

    return image_paths


def recognize_vat_invoice(access_token: str, image_path: str) -> dict:
    """调用百度OCR识别增值税发票"""
    url = f"https://aip.baidubce.com/rest/2.0/ocr/v1/vat_invoice?access_token={access_token}"

    image_base64 = image_to_base64(image_path)

    headers = {"Content-Type": "application/x-www-form-urlencoded"}
    data = {"image": image_base64}

    response = requests.post(url, headers=headers, data=data)
    return response.json()


def classify_invoice(commodity_name: str) -> str:
    """根据商品名称分类发票"""
    if not commodity_name:
        return "其他"

    commodity_lower = commodity_name.lower()

    for category, keywords in CATEGORY_RULES.items():
        for keyword in keywords:
            if keyword in commodity_lower:
                return category

    return "其他"


def parse_invoice_result(result: dict, file_path: str) -> dict:
    """解析OCR识别结果"""
    if "error_code" in result:
        print(f"识别失败 [{file_path}]: {result.get('error_msg', '未知错误')}")
        return None

    words_result = result.get("words_result", {})

    # 提取关键字段
    invoice_data = {
        "文件名": os.path.basename(file_path),
        "发票代码": words_result.get("InvoiceCode", ""),
        "发票号码": words_result.get("InvoiceNum", ""),
        "开票日期": words_result.get("InvoiceDate", ""),
        "销售方名称": words_result.get("SellerName", ""),
        "购买方名称": words_result.get("PurchaserName", ""),
        "商品名称": "",
        "金额": 0.0,
        "税额": 0.0,
        "价税合计": 0.0,
        "发票类型": words_result.get("InvoiceType", ""),
    }

    # 提取商品明细
    commodity_list = words_result.get("CommodityName", [])
    if isinstance(commodity_list, list) and commodity_list:
        invoice_data["商品名称"] = "; ".join([item.get("word", "") for item in commodity_list])
    elif isinstance(commodity_list, str):
        invoice_data["商品名称"] = commodity_list

    # 提取金额
    try:
        amount_str = words_result.get("TotalAmount", "0")
        if isinstance(amount_str, str):
            amount_str = amount_str.replace(",", "").replace("￥", "").replace("¥", "")
        invoice_data["金额"] = float(amount_str) if amount_str else 0.0
    except (ValueError, TypeError):
        invoice_data["金额"] = 0.0

    # 提取税额
    try:
        tax_str = words_result.get("TotalTax", "0")
        if isinstance(tax_str, str):
            tax_str = tax_str.replace(",", "").replace("￥", "").replace("¥", "")
        invoice_data["税额"] = float(tax_str) if tax_str else 0.0
    except (ValueError, TypeError):
        invoice_data["税额"] = 0.0

    # 提取价税合计
    try:
        total_str = words_result.get("AmountInFiguers", "0")
        if isinstance(total_str, str):
            total_str = total_str.replace(",", "").replace("￥", "").replace("¥", "")
        invoice_data["价税合计"] = float(total_str) if total_str else 0.0
    except (ValueError, TypeError):
        invoice_data["价税合计"] = invoice_data["金额"] + invoice_data["税额"]

    # 自动分类
    invoice_data["分类"] = classify_invoice(invoice_data["商品名称"])

    # 提取月份
    date_str = invoice_data["开票日期"]
    if date_str:
        try:
            # 尝试解析日期格式：2024年01月15日 或 2024-01-15
            date_str = date_str.replace("年", "-").replace("月", "-").replace("日", "")
            date_obj = datetime.strptime(date_str.strip(), "%Y-%m-%d")
            invoice_data["月份"] = date_obj.strftime("%Y-%m")
        except ValueError:
            invoice_data["月份"] = "未知"
    else:
        invoice_data["月份"] = "未知"

    return invoice_data


def get_invoice_files(directory: str) -> list:
    """获取目录下所有发票文件"""
    supported_extensions = {".jpg", ".jpeg", ".png", ".bmp", ".pdf"}
    files = []

    for file in Path(directory).iterdir():
        if file.is_file() and file.suffix.lower() in supported_extensions:
            files.append(str(file))

    return sorted(files)


def create_excel_report(invoices: list, output_path: str):
    """生成Excel报表"""
    wb = Workbook()

    # 样式定义
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # ========== Sheet 1: 发票明细 ==========
    ws1 = wb.active
    ws1.title = "发票明细"

    columns = ["文件名", "发票号码", "开票日期", "销售方名称", "商品名称", "金额", "税额", "价税合计", "分类", "月份"]

    # 写入表头
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws1.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写入数据
    for row_idx, inv in enumerate(invoices, 2):
        for col_idx, col_name in enumerate(columns, 1):
            value = inv.get(col_name, "")
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if isinstance(value, float):
                cell.number_format = "#,##0.00"

    # 调整列宽
    col_widths = [20, 15, 15, 25, 30, 12, 12, 12, 10, 10]
    for i, width in enumerate(col_widths, 1):
        ws1.column_dimensions[chr(64 + i)].width = width

    # ========== Sheet 2: 按类别汇总 ==========
    ws2 = wb.create_sheet("按类别汇总")

    category_summary = defaultdict(lambda: {"数量": 0, "金额": 0.0, "税额": 0.0, "价税合计": 0.0})
    for inv in invoices:
        cat = inv.get("分类", "其他")
        category_summary[cat]["数量"] += 1
        category_summary[cat]["金额"] += inv.get("金额", 0)
        category_summary[cat]["税额"] += inv.get("税额", 0)
        category_summary[cat]["价税合计"] += inv.get("价税合计", 0)

    headers2 = ["分类", "发票数量", "金额合计", "税额合计", "价税合计"]
    for col_idx, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    row_idx = 2
    total_count, total_amount, total_tax, total_sum = 0, 0.0, 0.0, 0.0
    for cat, data in sorted(category_summary.items()):
        ws2.cell(row=row_idx, column=1, value=cat).border = thin_border
        ws2.cell(row=row_idx, column=2, value=data["数量"]).border = thin_border
        ws2.cell(row=row_idx, column=3, value=data["金额"]).border = thin_border
        ws2.cell(row=row_idx, column=4, value=data["税额"]).border = thin_border
        ws2.cell(row=row_idx, column=5, value=data["价税合计"]).border = thin_border
        ws2.cell(row=row_idx, column=3).number_format = "#,##0.00"
        ws2.cell(row=row_idx, column=4).number_format = "#,##0.00"
        ws2.cell(row=row_idx, column=5).number_format = "#,##0.00"
        total_count += data["数量"]
        total_amount += data["金额"]
        total_tax += data["税额"]
        total_sum += data["价税合计"]
        row_idx += 1

    # 合计行
    ws2.cell(row=row_idx, column=1, value="合计").font = Font(bold=True)
    ws2.cell(row=row_idx, column=2, value=total_count).font = Font(bold=True)
    ws2.cell(row=row_idx, column=3, value=total_amount).font = Font(bold=True)
    ws2.cell(row=row_idx, column=4, value=total_tax).font = Font(bold=True)
    ws2.cell(row=row_idx, column=5, value=total_sum).font = Font(bold=True)
    for col in range(1, 6):
        ws2.cell(row=row_idx, column=col).border = thin_border
    ws2.cell(row=row_idx, column=3).number_format = "#,##0.00"
    ws2.cell(row=row_idx, column=4).number_format = "#,##0.00"
    ws2.cell(row=row_idx, column=5).number_format = "#,##0.00"

    # ========== Sheet 3: 按月份汇总 ==========
    ws3 = wb.create_sheet("按月份汇总")

    month_summary = defaultdict(lambda: {"数量": 0, "金额": 0.0, "税额": 0.0, "价税合计": 0.0})
    for inv in invoices:
        month = inv.get("月份", "未知")
        month_summary[month]["数量"] += 1
        month_summary[month]["金额"] += inv.get("金额", 0)
        month_summary[month]["税额"] += inv.get("税额", 0)
        month_summary[month]["价税合计"] += inv.get("价税合计", 0)

    headers3 = ["月份", "发票数量", "金额合计", "税额合计", "价税合计"]
    for col_idx, header in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    row_idx = 2
    for month, data in sorted(month_summary.items()):
        ws3.cell(row=row_idx, column=1, value=month).border = thin_border
        ws3.cell(row=row_idx, column=2, value=data["数量"]).border = thin_border
        ws3.cell(row=row_idx, column=3, value=data["金额"]).border = thin_border
        ws3.cell(row=row_idx, column=4, value=data["税额"]).border = thin_border
        ws3.cell(row=row_idx, column=5, value=data["价税合计"]).border = thin_border
        ws3.cell(row=row_idx, column=3).number_format = "#,##0.00"
        ws3.cell(row=row_idx, column=4).number_format = "#,##0.00"
        ws3.cell(row=row_idx, column=5).number_format = "#,##0.00"
        row_idx += 1

    # ========== Sheet 4: 报销清单 ==========
    ws4 = wb.create_sheet("报销清单")

    ws4.cell(row=1, column=1, value="报销汇总清单").font = Font(bold=True, size=14)
    ws4.merge_cells("A1:D1")

    ws4.cell(row=3, column=1, value="生成日期：")
    ws4.cell(row=3, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M"))

    ws4.cell(row=4, column=1, value="发票总数：")
    ws4.cell(row=4, column=2, value=len(invoices))

    ws4.cell(row=5, column=1, value="金额合计：")
    ws4.cell(row=5, column=2, value=total_amount)
    ws4.cell(row=5, column=2).number_format = "#,##0.00"

    ws4.cell(row=6, column=1, value="税额合计：")
    ws4.cell(row=6, column=2, value=total_tax)
    ws4.cell(row=6, column=2).number_format = "#,##0.00"

    ws4.cell(row=7, column=1, value="价税合计：")
    ws4.cell(row=7, column=2, value=total_sum)
    ws4.cell(row=7, column=2).number_format = "#,##0.00"

    # 日期范围
    dates = [inv.get("开票日期", "") for inv in invoices if inv.get("开票日期")]
    if dates:
        ws4.cell(row=8, column=1, value="日期范围：")
        ws4.cell(row=8, column=2, value=f"{min(dates)} 至 {max(dates)}")

    # 保存文件
    wb.save(output_path)
    print(f"\n报表已生成: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="发票自动归档工具")
    parser.add_argument("directory", nargs="?", default=".", help="发票文件所在目录（默认当前目录）")
    parser.add_argument("-o", "--output", default="发票汇总报表.xlsx", help="输出Excel文件名")
    args = parser.parse_args()

    # 加载配置
    config = load_config()

    # 获取access_token
    print("正在获取API访问令牌...")
    access_token = get_access_token(config["api_key"], config["secret_key"])

    # 获取发票文件列表
    invoice_files = get_invoice_files(args.directory)
    if not invoice_files:
        print(f"错误：目录 {args.directory} 下没有找到发票文件")
        sys.exit(1)

    print(f"找到 {len(invoice_files)} 个发票文件")

    # 处理每个发票
    invoices = []
    temp_files = []

    for i, file_path in enumerate(invoice_files, 1):
        print(f"[{i}/{len(invoice_files)}] 正在识别: {os.path.basename(file_path)}")

        # 处理PDF文件
        if file_path.lower().endswith(".pdf"):
            image_paths = pdf_to_images(file_path, args.directory)
            for img_path in image_paths:
                result = recognize_vat_invoice(access_token, img_path)
                invoice_data = parse_invoice_result(result, file_path)
                if invoice_data:
                    invoices.append(invoice_data)
                temp_files.append(img_path)
        else:
            result = recognize_vat_invoice(access_token, file_path)
            invoice_data = parse_invoice_result(result, file_path)
            if invoice_data:
                invoices.append(invoice_data)

    # 清理临时文件
    for temp_file in temp_files:
        try:
            os.remove(temp_file)
        except OSError:
            pass

    if not invoices:
        print("错误：没有成功识别任何发票")
        sys.exit(1)

    print(f"\n成功识别 {len(invoices)} 张发票")

    # 生成报表
    output_path = os.path.join(args.directory, args.output)
    create_excel_report(invoices, output_path)

    # 打印汇总信息
    total = sum(inv.get("价税合计", 0) for inv in invoices)
    print(f"\n===== 汇总 =====")
    print(f"发票总数: {len(invoices)} 张")
    print(f"价税合计: ¥{total:,.2f}")


if __name__ == "__main__":
    main()
